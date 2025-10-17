import base64
import io
import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook
from datetime import date
import calendar
from babel.dates import format_date

from odoo import models, fields, api
import pandas as pd

class XmlImport(models.Model):
    _name = "xml.import"
    _description = "XML Import to XLS"

    name = fields.Char(string="Name", compute="_compute_name", store=True)
    xml_file = fields.Binary("XML File", required=True)
    xml_filename = fields.Char("XML Filename")
    template_id = fields.Many2one("xml.xls.template", string="XLS Template")
    xls_file = fields.Binary("Filled XLS")
    xls_filename = fields.Char("Filled Filename")
    partner_id = fields.Many2one("res.partner", string="Partner")

    date_start = fields.Date("Start of Period")
    date_end = fields.Date("End of Period")
    date_prev_end = fields.Date("End of Period (Prev Year)")
    
    html_table = fields.Html("Preview Table", compute="_compute_html_table", sanitize=False)

    def action_print_report(self):
        self.ensure_one()
        return self.env.ref('xml_to_xls.action_report_sheetjs').report_action(self)

    def _compute_html_table(self):
        """Генерация HTML-таблицы на основе XLS-шаблона с подставленными данными"""
        for rec in self:
            rec.html_table = rec._generate_html_from_xls()

    def _decode_base64(self, data):
        """Декодирует base64, исправляя паддинг"""
        if not data:
            return None
        data_str = data.decode() if isinstance(data, bytes) else data
        missing_padding = len(data_str) % 4
        if missing_padding:
            data_str += "=" * (4 - missing_padding)
        return base64.b64decode(data_str) 

    def generate_html_from_xls(self):  
        """Генерация HTML-таблицы на основе XLS-шаблона с подставленными данными"""
        self._generate_html_from_xls()         

    def _generate_html_from_xls(self):
        self.ensure_one()
        # for rec in self:
        #     if not rec.xls_file:
        #         rec.html_table = "<p>No data</p>"
        #         continue

        #     # Загружаем Excel из бинарного поля
        #     # Декодируем base64 и создаем BytesIO
        #     try:
        #         excel_bytes = io.BytesIO(self._decode_base64(rec.xls_file))

        #     # Явно указываем движок openpyxl для XLSX
        #         df = pd.read_excel(excel_bytes, sheet_name=None, engine='openpyxl')
        #         html_parts = []
        #         for sheet_name, data in df.items():
        #             html_table = data.to_html(index=False)  # pandas генерирует HTML таблицу
        #             html_parts.append(f"<h3>{sheet_name}</h3>{html_table}")

        #         return ''.join(html_parts)
        #         # rec.html_table = html_table
        #         pass
        #     except Exception as e:
        #         return f"<p>Error reading XLSX: {e}</p>"
        if not self.xls_file:
            return "<em>No XLS generated</em>"

        try:
            xls_bytes = base64.b64decode(self.xls_file)
            wb = load_workbook(io.BytesIO(xls_bytes))
            ws = wb.active

            # html = ['<table style="border-collapse:collapse;width:100%;">']
            html = ['<table style="border: none; width:100%;">']
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                html.append("<tr>")
                for cell in row:
                    val = "" if cell is None else str(cell)
                    # html.append(f"<td style='border:1px solid #000;padding:4px;'>{val}</td>")
                    # --- жирное выделение для 1-й и 3-й строки ---
                    if row_idx in (1, 3):
                        html.append(f"<td style='border: none; padding:2px; font-weight:bold;'>{val}</td>")
                    else:
                        html.append(f"<td style='border: none; padding:2px;'>{val}</td>")
                html.append("</tr>")
            html.append("</table>")
            return "".join(html)

        except Exception as e:
            return f"<em>Error building HTML: {e}</em>"

    def name_get(self):
        res = []
        for rec in self:
            parts = []
            if rec.template_id:
                parts.append(rec.template_id.name)
            if rec.date_end:
                parts.append(format_date(rec.date_end, format='d MMMM y', locale='en'))
            if rec.partner_id:
                parts.append(rec.partner_id.name)

            display = " — ".join(parts) if parts else (rec.name or "XML Import")
            res.append((rec.id, display))
        return res

    # --- общая логика формирования name ---
    @api.depends("template_id", "date_end")
    def _compute_name(self):
        for rec in self:
            if rec.template_id and rec.date_end:
                rec.name = f"{rec.template_id.name} on {format_date(rec.date_end, format='d MMMM y', locale='en')}"
            elif rec.template_id:
                rec.name = rec.template_id.name
            else:
                rec.name = "XML Import" 

    @api.onchange("xml_file")
    def _compute_period_dates(self):
        for rec in self:
            rec.date_start = rec.date_end = rec.date_prev_end = False
            if not rec.xml_file:
                continue

            try:
                xml_content = base64.b64decode(rec.xml_file)
                root = ET.fromstring(xml_content)

                period_month = int(root.findtext(".//PERIOD_MONTH") or 0)
                period_type = int(root.findtext(".//PERIOD_TYPE") or 0)
                period_year = int(root.findtext(".//PERIOD_YEAR") or 0)

                if not (period_month and period_type and period_year):
                    continue

                # 1. Начало года
                start_year = date(period_year, 1, 1)

                # 2. Конец периода
                if period_type == 1:  # месяц
                    last_day = calendar.monthrange(period_year, period_month)[1]
                    end_period = date(period_year, period_month, last_day)
                elif period_type == 2:  # квартал
                    quarter_end_month = ((period_month - 1) // 3 + 1) * 3
                    last_day = calendar.monthrange(period_year, quarter_end_month)[1]
                    end_period = date(period_year, quarter_end_month, last_day)
                elif period_type == 3:  # полугодие
                    if period_month <= 6:
                        last_day = calendar.monthrange(period_year, 6)[1]
                        end_period = date(period_year, 6, last_day)
                    else:
                        last_day = calendar.monthrange(period_year, 12)[1]
                        end_period = date(period_year, 12, last_day)
                elif period_type == 4:  # год
                    last_day = calendar.monthrange(period_year, 12)[1]
                    end_period = date(period_year, 12, last_day)
                else:
                    last_day = calendar.monthrange(period_year, period_month)[1]
                    end_period = date(period_year, period_month, last_day)

                # 3. Конец периода год назад
                try:
                    end_period_prev = end_period.replace(year=end_period.year - 1)
                except ValueError:
                    end_period_prev = end_period.replace(year=end_period.year - 1, day=28)

                rec.date_start = start_year
                rec.date_end = end_period
                rec.date_prev_end = end_period_prev

            except Exception:
                continue

    @api.onchange('xml_file')
    def onchange_xml_file(self):
        for rec in self:
            if rec.xml_file:
                # Декодируем XML
                xml_content = base64.b64decode(rec.xml_file)
                root = ET.fromstring(xml_content)

                # Примитивное определение типа отчета по C_DOC_SUB в DECLARHEAD
                report_type = self._detect_type_from_xml(root)
                if report_type:
                    template = self.env['xml.xls.template'].search([('report_type', '=', report_type)], limit=1)
                    rec.template_id = template.id if template else False
                else:
                    rec.template_id = False

                rec.partner_id = self._detect_partner_from_xml(root)  # Сброс партнера при смене файла    

    def _detect_partner_from_xml(self, root):
        """
        Примитивное определение по C_DOC_SUB в DECLARHEAD if present.
        Возвращает 'balance' for C_DOC_SUB == '001', 'profit_loss' for '002', else None.
        """
        try:
            node = root.find('.//FIRM_EDRPOU')
            if node is not None and node.text:
                code = node.text.strip()
                partner = self.env['res.partner'].search([('okpo_code', '=', code)], limit=1)
                if partner:
                    return partner.id
                # доп. правила можно добавить
        except Exception:
            pass
        return False                

    def _detect_type_from_xml(self, root):
        """
        Примитивное определение по C_DOC_SUB в DECLARHEAD if present.
        Возвращает 'balance' for C_DOC_SUB == '001', 'profit_loss' for '002', else None.
        """
        try:
            node = root.find('.//C_DOC_SUB')
            if node is not None and node.text:
                code = node.text.strip()
                if code == '001':
                    return 'balance'
                if code == '002':
                    return 'profit_loss'
                # доп. правила можно добавить
        except Exception:
            pass
        return None

    def action_preview_xls(self):
        return True    
        # 1. Начало года
        start_year = date(period_year, 1, 1)

        # 2. Конец периода
        if period_type == 1:  # месяц
            last_day = calendar.monthrange(period_year, period_month)[1]
            end_period = date(period_year, period_month, last_day)
        elif period_type == 2:  # квартал
            quarter_end_month = ((period_month - 1)//3 + 1) * 3
            last_day = calendar.monthrange(period_year, quarter_end_month)[1]
            end_period = date(period_year, quarter_end_month, last_day)
        elif period_type == 3:  # полугодие
            if period_month <= 6:
                last_day = calendar.monthrange(period_year, 6)[1]
                end_period = date(period_year, 6, last_day)
            else:
                last_day = calendar.monthrange(period_year, 12)[1]
                end_period = date(period_year, 12, last_day)
        elif period_type == 4:  # год
            last_day = calendar.monthrange(period_year, 12)[1]
            end_period = date(period_year, 12, last_day)
        else:
            # fallback: конец месяца
            last_day = calendar.monthrange(period_year, period_month)[1]
            end_period = date(period_year, period_month, last_day)

        # 3. Конец периода год назад
        try:
            end_period_prev = end_period.replace(year=end_period.year - 1)
        except ValueError:
            # для 29 февраля
            end_period_prev = end_period.replace(year=end_period.year - 1, day=28)

        # Форматирование
        fmt = "%d %B %Y"
        return start_year.strftime(fmt), end_period.strftime(fmt), end_period_prev.strftime(fmt)    

    def action_fill_template(self):
        # fmt = "%d %B %Y"
        for rec in self:
            rec._compute_period_dates()
            # Загружаем XML
            xml_content = base64.b64decode(rec.xml_file)
            root = ET.fromstring(xml_content)

            # --- страховка: если partner_id или template_id пустые ---
            if not rec.partner_id:
                partner_id = rec._detect_partner_from_xml(root)
                if partner_id:
                    rec.partner_id = partner_id

            if not rec.template_id:
                report_type = rec._detect_type_from_xml(root)
                if report_type:
                    template = self.env['xml.xls.template'].search([('report_type', '=', report_type)], limit=1)
                    if template:
                        rec.template_id = template


            # Собираем все значения в словарь {tag: text}
            values = {elem.tag: elem.text for elem in root.iter() if elem.text}

            # Загружаем XLSX шаблон
            template_bytes = base64.b64decode(rec.template_id.template_file)
            wb = load_workbook(io.BytesIO(template_bytes))

            # Проходим по всем листам и ячейкам
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            tag = cell.value.strip()
                            if tag in values:
                                val = values[tag]

                                # --- спец. правила для A/B тегов ---
                                if re.match(r'^[AB]\d+$', tag):
                                    try:
                                        num = float(val)
                                        if num == 0:
                                            cell.value = "-"
                                        elif num < 0:
                                            cell.value = f"({abs(int(num) if num.is_integer() else num)})"
                                        else:
                                            cell.value = int(num) if num.is_integer() else num
                                    except Exception:
                                        # если не число — оставляем как есть
                                        cell.value = val
                                else:
                                    cell.value = val
                                continue    
                            if tag == 'Partner' and rec.partner_id:
                                cell.value = rec.partner_id.name_eng or rec.partner_id.name
                            if tag == 'PartnerAddress' and rec.partner_id:
                                cell.value = rec.partner_id.address_eng or rec.partner_id.contact_address
                            if re.match(r'^[AB]\d+$', tag):
                                cell.value = "-"  # если тега нет в XML, ставим прочерк  
                            if tag == 'DateStart' and rec.date_start:
                                cell.value = format_date(rec.date_start, format="d MMMM y", locale="en")
                            if tag == 'DateEnd' and rec.date_end:
                                cell.value = format_date(rec.date_end, format="d MMMM y", locale="en")
                            if tag == 'DatePrevEnd' and rec.date_prev_end:
                                cell.value = format_date(rec.date_prev_end, format="d MMMM y", locale="en")              

            # Сохраняем результат
            out = io.BytesIO()
            wb.save(out)
            rec.xls_file = base64.b64encode(out.getvalue())
            filename = rec.xml_filename
            if filename.lower().endswith(".xml"):
                filename = filename[:-4] + ".xlsx"
            rec.xls_filename = f"{rec.template_id.report_type}_{filename}"